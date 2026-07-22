from __future__ import annotations

import calendar
import io
import math
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.datetime import from_excel

APP_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = APP_DIR / "output_template.xlsx"

AFF_CV_CELLS = ["W7", "AI7", "AU7", "BG7", "BS7"]
AFF_COST_CELLS = ["Y7", "AK7", "AW7", "BI7", "BU7"]
DISPLAY_ROWS = [32, 33, 34, 35]

OPERATIONAL_MEDIA_DETAILS = {
    "検索": [
        "GAD_GEN_[01.公営競技]_C001",
        "GAD_GEN_[04.円預金]_C002",
        "GAD_GEN_[06.カード]_C003",
        "GAD_GEN_[デビットカード切り出し]_C004",
        "GAD_GEN_[円預金.子供口座]_C005",
        "YSS_GEN_[01.公営競技]_STD_C001",
        "YSS_GEN_[04.円預金]_STD_C002",
        "YSS_GEN_[06.カード]_STD_C003",
        "YSS_GEN_[円預金.子供口座]_STD_C005",
        "YSS_GEN_[デビットカード切り出し]_STD_C004_[in_bank_4]",
        "MSA_GEN_[04.円預金]_C002",
        "MSA_GEN_[円預金.子供口座]_C005",
        "MSA_GEN_[06.カード]_C003",
        "MSA_GEN_[01.公営競技]_C001",
        "配信開始後記載⑨",
    ],
    "ディスプレイ": [
        "GDN_[04.円預金]_C002",
        "GDN_[06.カード]_C002",
        "LYDA_[04.円預金]_C001_[in_bank_15]",
        "META_[04.円預金]",
        "META_[06.カード]",
        "SmartNews_[04.円預金]",
        "X_[04.円預金]_C001",
        "YouTube_スキップ不可_[in_bank_3]",
    ],
}

OPERATIONAL_RAW_SHEETS = [
    "【非表示】GDNローデータ",
    "【非表示】Metaローデータ",
    "【非表示】SmartNewsローデータ",
    "【非表示】Xローデータ",
    "【非表示】YDAローデータ",
    "【非表示】MSAローデータ",
    "【非表示】Yahooローデータ",
    "【非表示】YouTubeローデータ",
    "【非表示】Googleローデータ",
]


@dataclass
class Metrics:
    cv: float = 0.0
    cost: float = 0.0

    @property
    def cpa(self) -> float:
        return safe_div(self.cost, self.cv)


def safe_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    text = str(value).strip().replace(",", "").replace("¥", "").replace("￥", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def safe_div(numerator: float, denominator: float) -> float:
    return numerator / denominator if denominator not in (0, 0.0) else 0.0


def normalize_text(value: Any) -> str:
    return str(value or "").replace("\u3000", " ").strip()


def normalize_datetime(value: Any, epoch=None) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, epoch=epoch)
        except Exception:
            return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"


class FastXlsxReader:
    """Reads cached Excel values directly from the XLSX ZIP, including hidden sheets."""
    def __init__(self, file_bytes: bytes):
        self._bytes = file_bytes
        self._zip = zipfile.ZipFile(io.BytesIO(file_bytes))
        self.shared_strings = self._read_shared_strings()
        self.sheets = self._read_sheet_map()
        self._cache: dict[str, dict[str, Any]] = {}

    def _read_shared_strings(self) -> list[str]:
        if "xl/sharedStrings.xml" not in self._zip.namelist():
            return []
        root = ET.fromstring(self._zip.read("xl/sharedStrings.xml"))
        return ["".join(t.text or "" for t in si.iter(f"{{{NS_MAIN}}}t")) for si in root.findall(f"{{{NS_MAIN}}}si")]

    def _read_sheet_map(self) -> dict[str, dict[str, str]]:
        workbook = ET.fromstring(self._zip.read("xl/workbook.xml"))
        rels = ET.fromstring(self._zip.read("xl/_rels/workbook.xml.rels"))
        rel_map = {r.attrib["Id"]: r.attrib["Target"] for r in rels.findall(f"{{{NS_PKG}}}Relationship")}
        result = {}
        for s in workbook.find(f"{{{NS_MAIN}}}sheets"):
            target = rel_map[s.attrib[f"{{{NS_REL}}}id"]]
            if not target.startswith("xl/"):
                target = "xl/" + target.lstrip("/")
            result[s.attrib["name"]] = {"state": s.attrib.get("state", "visible"), "target": target}
        return result

    @property
    def sheetnames(self) -> list[str]:
        return list(self.sheets.keys())

    def sheet_state(self, sheet_name: str) -> str:
        return self.sheets[sheet_name]["state"]

    def values(self, sheet_name: str) -> dict[str, Any]:
        if sheet_name in self._cache:
            return self._cache[sheet_name]
        root = ET.fromstring(self._zip.read(self.sheets[sheet_name]["target"]))
        values: dict[str, Any] = {}
        for cell in root.iter(f"{{{NS_MAIN}}}c"):
            ref = cell.attrib.get("r")
            if not ref:
                continue
            cell_type = cell.attrib.get("t")
            value_node = cell.find(f"{{{NS_MAIN}}}v")
            value: Any = None
            if cell_type == "s" and value_node is not None:
                idx = int(value_node.text or 0)
                value = self.shared_strings[idx] if idx < len(self.shared_strings) else None
            elif cell_type == "inlineStr":
                value = "".join(t.text or "" for t in cell.iter(f"{{{NS_MAIN}}}t"))
            elif value_node is not None:
                raw = value_node.text
                try:
                    value = float(raw) if raw is not None and ("." in raw or "E" in raw.upper()) else int(raw)
                except (ValueError, AttributeError):
                    value = raw
            values[ref] = value
        self._cache[sheet_name] = values
        return values

    def cell(self, sheet_name: str, address: str) -> Any:
        return self.values(sheet_name).get(address)


@st.cache_resource(show_spinner=False)
def load_fast_xlsx(file_bytes: bytes):
    return FastXlsxReader(file_bytes)


def selectable_sheets(file_bytes: bytes) -> list[str]:
    return load_fast_xlsx(file_bytes).sheetnames


def sheet_state(file_bytes: bytes, sheet_name: str) -> str:
    return load_fast_xlsx(file_bytes).sheet_state(sheet_name)

def get_cached_cell(reader: FastXlsxReader, sheet_name: str, address: str) -> float:
    return safe_number(reader.cell(sheet_name, address))


def read_aff_plan(file_bytes: bytes, sheet_name: str) -> tuple[Metrics, set[str]]:
    reader = load_fast_xlsx(file_bytes)
    vals = reader.values(sheet_name)
    cv = sum(get_cached_cell(reader, sheet_name, cell) for cell in AFF_CV_CELLS)
    cost = sum(get_cached_cell(reader, sheet_name, cell) for cell in AFF_COST_CELLS)
    sites = {normalize_text(v) for ref, v in vals.items() if ref[0] == "A" and ref[1:].isdigit() and normalize_text(v)}
    return Metrics(cv=cv, cost=cost), sites


def read_operational_plan(file_bytes: bytes, sheet_name: str) -> tuple[Metrics, Metrics]:
    reader = load_fast_xlsx(file_bytes)
    search = Metrics(cv=get_cached_cell(reader, sheet_name, "J11"), cost=get_cached_cell(reader, sheet_name, "I11"))
    display = Metrics(
        cv=sum(get_cached_cell(reader, sheet_name, f"J{r}") for r in DISPLAY_ROWS),
        cost=sum(get_cached_cell(reader, sheet_name, f"I{r}") for r in DISPLAY_ROWS),
    )
    return search, display

def read_aff_actual(file_bytes: bytes, valid_sites: set[str]) -> tuple[Metrics, pd.DataFrame]:
    # Rakuten CSV is usually CP932. Fall back safely for UTF-8 exports.
    df = None
    for encoding in ("cp932", "utf-8-sig", "utf-8"):
        try:
            df = pd.read_csv(io.BytesIO(file_bytes), encoding=encoding, dtype=str)
            break
        except UnicodeDecodeError:
            continue
    if df is None:
        raise ValueError("AFF実績CSVの文字コードを判定できませんでした。")

    required = ["パートナーサイト名", "件数", "成果発生日時", "グロス"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"AFF実績に必要な列がありません: {', '.join(missing)}")

    df["_site"] = df["パートナーサイト名"].map(normalize_text)
    df["_cv"] = pd.to_numeric(df["件数"].str.replace(",", "", regex=False), errors="coerce").fillna(0)
    df["_cost"] = pd.to_numeric(df["グロス"].str.replace(",", "", regex=False), errors="coerce").fillna(0)
    df["_date"] = pd.to_datetime(df["成果発生日時"], errors="coerce").dt.date
    matched = df[df["_site"].isin(valid_sites)].copy()
    actual = Metrics(cv=float(matched["_cv"].sum()), cost=float(matched["_cost"].sum()))
    return actual, matched


def find_sheet_name(reader: FastXlsxReader, wanted: str) -> str:
    candidates = [wanted, f"【{wanted}】"]
    for name in candidates:
        if name in reader.sheetnames:
            return name
    compact = wanted.replace("【", "").replace("】", "")
    for name in reader.sheetnames:
        if name.replace("【", "").replace("】", "") == compact:
            return name
    raise KeyError(f"シート『{wanted}』が見つかりません。")


def excel_serial_to_datetime(value: Any) -> datetime | None:
    if isinstance(value, (int, float)):
        try:
            return from_excel(value)
        except Exception:
            return None
    return normalize_datetime(value)


def detect_daily_date_column(vals: dict[str, Any], start_row: int = 23, end_row: int = 60) -> int:
    for col_num, col_letter in ((1, "A"), (2, "B")):
        found = sum(1 for row in range(start_row, end_row + 1) if excel_serial_to_datetime(vals.get(f"{col_letter}{row}")))
        if found >= 2:
            return col_num
    raise ValueError("日付列（A列またはB列）を判定できませんでした。")


def read_operational_actual(file_bytes: bytes) -> tuple[Metrics, Metrics, pd.DataFrame, pd.DataFrame]:
    reader = load_fast_xlsx(file_bytes)
    search_name = find_sheet_name(reader, "Search_合計")
    display_name = find_sheet_name(reader, "Display_合計")
    search_vals = reader.values(search_name)
    display_vals = reader.values(display_name)

    search = Metrics(cv=get_cached_cell(reader, search_name, "M22"), cost=get_cached_cell(reader, search_name, "I22"))
    display = Metrics(cv=get_cached_cell(reader, display_name, "M22"), cost=get_cached_cell(reader, display_name, "I22"))

    def daily(vals: dict[str, Any]) -> pd.DataFrame:
        date_col_num = detect_daily_date_column(vals)
        date_col = "A" if date_col_num == 1 else "B"
        rows = []
        for row in range(23, 100):
            dt = excel_serial_to_datetime(vals.get(f"{date_col}{row}"))
            if not dt:
                continue
            cv = safe_number(vals.get(f"M{row}"))
            cost = safe_number(vals.get(f"I{row}"))
            rows.append({"日付": dt.date(), "CV": cv, "コスト": cost, "CPA": safe_div(cost, cv)})
        return pd.DataFrame(rows)

    return search, display, daily(search_vals), daily(display_vals)

def infer_month(search_daily: pd.DataFrame, display_daily: pd.DataFrame, aff_matched: pd.DataFrame) -> tuple[int, int]:
    dates: list[date] = []
    for df in (search_daily, display_daily):
        if not df.empty:
            dates.extend(df["日付"].dropna().tolist())
    if not aff_matched.empty:
        dates.extend(aff_matched["_date"].dropna().tolist())
    if not dates:
        today = date.today()
        return today.year, today.month
    # The operational report contains all calendar dates, so use its earliest date.
    d = min(dates)
    return d.year, d.month


def build_summary(search_plan, display_plan, aff_plan, search_actual, display_actual, aff_actual) -> pd.DataFrame:
    rows = []
    for label, target, actual in [
        ("検索", search_plan, search_actual),
        ("ディスプレイ", display_plan, display_actual),
        ("アフィリエイト", aff_plan, aff_actual),
    ]:
        rows.append({
            "区分": label,
            "CV 目標": target.cv,
            "CV 実績": actual.cv,
            "CV TVA": safe_div(target.cv, actual.cv),
            "コスト 目標": target.cost,
            "コスト 実績": actual.cost,
            "コスト TVA": safe_div(target.cost, actual.cost),
            "CPA 目標": target.cpa,
            "CPA 実績": actual.cpa,
            "CPA TVA": safe_div(target.cpa, actual.cpa),
        })
    total_target = Metrics(cv=sum(r[1].cv for r in [("", search_plan, None), ("", display_plan, None), ("", aff_plan, None)]),
                           cost=search_plan.cost + display_plan.cost + aff_plan.cost)
    total_actual = Metrics(cv=search_actual.cv + display_actual.cv + aff_actual.cv,
                           cost=search_actual.cost + display_actual.cost + aff_actual.cost)
    rows.append({
        "区分": "合計",
        "CV 目標": total_target.cv, "CV 実績": total_actual.cv, "CV TVA": safe_div(total_target.cv, total_actual.cv),
        "コスト 目標": total_target.cost, "コスト 実績": total_actual.cost, "コスト TVA": safe_div(total_target.cost, total_actual.cost),
        "CPA 目標": total_target.cpa, "CPA 実績": total_actual.cpa, "CPA TVA": safe_div(total_target.cpa, total_actual.cpa),
    })
    return pd.DataFrame(rows)


def build_daily(year, month, search_plan, display_plan, aff_plan, search_daily, display_daily, aff_matched) -> pd.DataFrame:
    days = calendar.monthrange(year, month)[1]
    dates = [date(year, month, d) for d in range(1, days + 1)]
    base = pd.DataFrame({"日付": dates})

    def merge_actual(prefix: str, df: pd.DataFrame):
        nonlocal base
        if df.empty:
            temp = pd.DataFrame({"日付": dates, f"実績 {prefix} CV": 0.0, f"実績 {prefix} コスト": 0.0})
        else:
            temp = df[["日付", "CV", "コスト"]].copy()
            temp = temp.rename(columns={"CV": f"実績 {prefix} CV", "コスト": f"実績 {prefix} コスト"})
        base = base.merge(temp, on="日付", how="left")
        base[f"実績 {prefix} CV"] = base[f"実績 {prefix} CV"].fillna(0)
        base[f"実績 {prefix} コスト"] = base[f"実績 {prefix} コスト"].fillna(0)
        base[f"実績 {prefix} CPA"] = base.apply(lambda r: safe_div(r[f"実績 {prefix} コスト"], r[f"実績 {prefix} CV"]), axis=1)

    merge_actual("検索", search_daily)
    merge_actual("ディスプレイ", display_daily)

    aff_daily = (aff_matched.groupby("_date", as_index=False)[["_cv", "_cost"]].sum()
                 .rename(columns={"_date": "日付", "_cv": "CV", "_cost": "コスト"})) if not aff_matched.empty else pd.DataFrame()
    merge_actual("アフィリエイト", aff_daily)

    for prefix, target in [("検索", search_plan), ("ディスプレイ", display_plan), ("アフィリエイト", aff_plan)]:
        base[f"目標 {prefix} CV"] = target.cv / days
        base[f"目標 {prefix} コスト"] = target.cost / days
        base[f"目標 {prefix} CPA"] = safe_div(target.cost, target.cv)
        for metric in ("CV", "コスト", "CPA"):
            base[f"目標vs実績 {prefix} {metric}"] = base.apply(
                lambda r, p=prefix, m=metric: safe_div(r[f"目標 {p} {m}"], r[f"実績 {p} {m}"]), axis=1
            )

    for kind in ("実績", "目標", "目標vs実績"):
        if kind == "目標vs実績":
            # Total target vs total actual, not a sum of percentages.
            base[f"{kind} 合計 CV"] = base.apply(lambda r: safe_div(sum(r[f"目標 {p} CV"] for p in ("検索","ディスプレイ","アフィリエイト")), sum(r[f"実績 {p} CV"] for p in ("検索","ディスプレイ","アフィリエイト"))), axis=1)
            base[f"{kind} 合計 コスト"] = base.apply(lambda r: safe_div(sum(r[f"目標 {p} コスト"] for p in ("検索","ディスプレイ","アフィリエイト")), sum(r[f"実績 {p} コスト"] for p in ("検索","ディスプレイ","アフィリエイト"))), axis=1)
            base[f"{kind} 合計 CPA"] = base.apply(lambda r: safe_div(safe_div(sum(r[f"目標 {p} コスト"] for p in ("検索","ディスプレイ","アフィリエイト")), sum(r[f"目標 {p} CV"] for p in ("検索","ディスプレイ","アフィリエイト"))), safe_div(sum(r[f"実績 {p} コスト"] for p in ("検索","ディスプレイ","アフィリエイト")), sum(r[f"実績 {p} CV"] for p in ("検索","ディスプレイ","アフィリエイト")))), axis=1)
        else:
            base[f"{kind} 合計 CV"] = sum(base[f"{kind} {p} CV"] for p in ("検索","ディスプレイ","アフィリエイト"))
            base[f"{kind} 合計 コスト"] = sum(base[f"{kind} {p} コスト"] for p in ("検索","ディスプレイ","アフィリエイト"))
            base[f"{kind} 合計 CPA"] = base.apply(lambda r, k=kind: safe_div(r[f"{k} 合計 コスト"], r[f"{k} 合計 CV"]), axis=1)

    order = ["日付"]
    for kind in ("実績", "目標", "目標vs実績"):
        for prefix in ("検索", "ディスプレイ", "アフィリエイト", "合計"):
            order += [f"{kind} {prefix} CV", f"{kind} {prefix} コスト", f"{kind} {prefix} CPA"]
    return base[order]


def canonical_campaign_name(value: Any) -> str:
    text = normalize_text(value)
    # Monthly raw files may append a tracking suffix such as _[in_bank_3].
    return re.sub(r"_?\[in_bank_\d+\]$", "", text, flags=re.IGNORECASE)


def campaign_matches(actual: Any, target: str) -> bool:
    actual_name = canonical_campaign_name(actual)
    target_name = canonical_campaign_name(target)
    return actual_name == target_name or actual_name.startswith(target_name + "_")


def read_operational_media(file_bytes: bytes) -> pd.DataFrame:
    reader = load_fast_xlsx(file_bytes)
    targets = [(category, detail) for category, details in OPERATIONAL_MEDIA_DETAILS.items() for detail in details]
    totals = {normalize_text(detail): {"コスト": 0.0, "CV": 0.0} for _, detail in targets}

    available_sheets = [name for name in OPERATIONAL_RAW_SHEETS if name in reader.sheetnames]
    if not available_sheets:
        raise KeyError("運用型実績に対象のローデータシートが見つかりません。")

    for sheet_name in available_sheets:
        vals = reader.values(sheet_name)
        for ref, raw_name in vals.items():
            if not (ref.startswith("A") and ref[1:].isdigit()):
                continue
            matched_target = next((detail for _, detail in targets if campaign_matches(raw_name, detail)), None)
            if matched_target is None:
                continue
            row_num = ref[1:]
            key = normalize_text(matched_target)
            totals[key]["コスト"] += safe_number(vals.get(f"G{row_num}"))
            totals[key]["CV"] += safe_number(vals.get(f"H{row_num}"))

    rows = []
    for category, details in OPERATIONAL_MEDIA_DETAILS.items():
        for detail in details:
            normalized = normalize_text(detail)
            cost = totals[normalized]["コスト"]
            cv = totals[normalized]["CV"]
            media_name = detail.split("_", 1)[0]
            rows.append({
                "媒体カテゴリ": category,
                "媒体名称": media_name,
                "媒体詳細": detail,
                "コスト": cost,
                "CV": cv,
                "CPA": safe_div(cost, cv),
            })
    return pd.DataFrame(rows)


def build_media(aff_matched: pd.DataFrame, operational_media: pd.DataFrame) -> pd.DataFrame:
    rows = operational_media.to_dict("records")
    if not aff_matched.empty:
        grouped = aff_matched.groupby("_site", as_index=False)[["_cost", "_cv"]].sum()
        for _, r in grouped.iterrows():
            rows.append({"媒体カテゴリ": "AFF", "媒体名称": r["_site"], "媒体詳細": "", "コスト": r["_cost"], "CV": r["_cv"], "CPA": safe_div(r["_cost"], r["_cv"])})
    return pd.DataFrame(rows)


def apply_table_style(ws, max_row: int, max_col: int):
    thin = Side(style="thin", color="D9E1F2")
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A4" if ws.title == "日別" else "A3"



def apply_workbook_font(wb, font_name: str = "Meiryo UI"):
    """
    ワークブック内すべてのセルを Meiryo UI に変更する。
    既存の太字・サイズ・色などは維持する。
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                new_font = copy(cell.font)
                new_font.name = font_name
                new_font.scheme = None
                cell.font = new_font

def export_excel(summary: pd.DataFrame, daily: pd.DataFrame, media: pd.DataFrame) -> bytes:
    if TEMPLATE_PATH.exists():
        wb = load_workbook(TEMPLATE_PATH)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.title = "合算"
        wb.create_sheet("日別")
        wb.create_sheet("媒体別")

    # 合算
    ws = wb["合算"]
    for r in range(3, 7):
        row = summary.iloc[r - 3]
        ws.cell(r, 1, row["区分"])
        values = [row["CV 目標"], row["CV 実績"], row["CV TVA"], row["コスト 目標"], row["コスト 実績"], row["コスト TVA"], row["CPA 目標"], row["CPA 実績"], row["CPA TVA"]]
        for c, value in enumerate(values, 2):
            ws.cell(r, c, float(value))
    for c in (2,3,5,6,8,9): ws.cell(3,c).number_format = ws.cell(3,c).number_format or '#,##0.00'
    for col in (4,7,10):
        for r in range(3,7): ws.cell(r,col).number_format = '0.0%'
    for col in (2,3):
        for r in range(3,7): ws.cell(r,col).number_format = '#,##0.00'
    for col in (5,6,8,9):
        for r in range(3,7): ws.cell(r,col).number_format = '¥#,##0'

    # 日別: template has 3 header rows, write row 4 onward.
    ws = wb["日別"]
    for row_idx, row in daily.iterrows():
        excel_row = row_idx + 4
        ws.cell(excel_row, 1, row["日付"])
        ws.cell(excel_row, 1).number_format = 'm/d'
        col = 2
        for kind in ("実績", "目標", "目標vs実績"):
            for prefix in ("検索", "ディスプレイ", "アフィリエイト", "合計"):
                for metric in ("CV", "コスト", "CPA"):
                    value = row[f"{kind} {prefix} {metric}"]
                    ws.cell(excel_row, col, float(value))
                    ws.cell(excel_row, col).number_format = '0.0%' if kind == "目標vs実績" else ('¥#,##0' if metric in ("コスト", "CPA") else '#,##0.00')
                    col += 1
    # clear any template date rows beyond selected month
    for r in range(4 + len(daily), ws.max_row + 1):
        for c in range(1, 38): ws.cell(r,c).value = None

    # 媒体別
    ws = wb["媒体別"]
    headers = ["媒体カテゴリ", "媒体名称", "媒体詳細", "コスト", "CV", "CPA"]
    for c,h in enumerate(headers,1): ws.cell(1,c,h)
    for r in range(2, ws.max_row + 1):
        for c in range(1,7): ws.cell(r,c).value = None
    for idx, row in media.iterrows():
        r = idx + 2
        for c,h in enumerate(headers,1): ws.cell(r,c,row[h])
        ws.cell(r,4).number_format='¥#,##0'; ws.cell(r,5).number_format='#,##0.00'; ws.cell(r,6).number_format='¥#,##0'

    apply_workbook_font(wb)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def display_summary(df: pd.DataFrame):
    styled = df.copy()
    st.dataframe(
        styled,
        use_container_width=True,
        hide_index=True,
        column_config={
            "CV TVA": st.column_config.NumberColumn(format="%.1f%%"),
            "コスト TVA": st.column_config.NumberColumn(format="%.1f%%"),
            "CPA TVA": st.column_config.NumberColumn(format="%.1f%%"),
            "コスト 目標": st.column_config.NumberColumn(format="¥%,.0f"),
            "コスト 実績": st.column_config.NumberColumn(format="¥%,.0f"),
            "CPA 目標": st.column_config.NumberColumn(format="¥%,.0f"),
            "CPA 実績": st.column_config.NumberColumn(format="¥%,.0f"),
        },
    )


def main():
    st.set_page_config(page_title="統合レポート作成", layout="wide")
    st.title("統合レポート作成")
    st.caption("AFF・運用型広告のプランと実績を結合し、合算／日別／媒体別レポートを作成します。")

    c1, c2 = st.columns(2)
    with c1:
        st.subheader("① AFF")
        aff_plan_file = st.file_uploader("AFFプラン（Excel）", type=["xlsx", "xlsm"], key="aff_plan")
        aff_actual_file = st.file_uploader("AFF実績（CSV）", type=["csv"], key="aff_actual")
        aff_sheet = None
        if aff_plan_file:
            aff_bytes = aff_plan_file.getvalue()
            sheets = selectable_sheets(aff_bytes)
            aff_sheet = st.selectbox("AFFプランのシート", sheets, key="aff_sheet")
            if sheet_state(aff_bytes, aff_sheet) != "visible":
                st.info("非表示シートを選択中です。非表示のまま読み込めます。")

    with c2:
        st.subheader("② 運用型")
        op_plan_file = st.file_uploader("運用型プラン（Excel）", type=["xlsx", "xlsm"], key="op_plan")
        op_actual_file = st.file_uploader("運用型実績（Excel）", type=["xlsx", "xlsm"], key="op_actual")
        op_sheet = None
        if op_plan_file:
            op_bytes = op_plan_file.getvalue()
            sheets = selectable_sheets(op_bytes)
            op_sheet = st.selectbox("運用型プランのシート", sheets, key="op_sheet")
            if sheet_state(op_bytes, op_sheet) != "visible":
                st.info("非表示シートを選択中です。非表示のまま読み込めます。")

    ready = all([aff_plan_file, aff_actual_file, op_plan_file, op_actual_file, aff_sheet, op_sheet])
    if not ready:
        st.info("4ファイルをアップロードし、プランシートを選択してください。")
        return

    if st.button("統合レポートを作成", type="primary", use_container_width=True):
        try:
            with st.spinner("集計しています…"):
                aff_plan, aff_sites = read_aff_plan(aff_plan_file.getvalue(), aff_sheet)
                search_plan, display_plan = read_operational_plan(op_plan_file.getvalue(), op_sheet)
                aff_actual, aff_matched = read_aff_actual(aff_actual_file.getvalue(), aff_sites)
                search_actual, display_actual, search_daily, display_daily = read_operational_actual(op_actual_file.getvalue())
                operational_media = read_operational_media(op_actual_file.getvalue())
                year, month = infer_month(search_daily, display_daily, aff_matched)
                summary = build_summary(search_plan, display_plan, aff_plan, search_actual, display_actual, aff_actual)
                daily = build_daily(year, month, search_plan, display_plan, aff_plan, search_daily, display_daily, aff_matched)
                media = build_media(aff_matched, operational_media)
                excel_bytes = export_excel(summary, daily, media)

            st.session_state["result"] = (summary, daily, media, excel_bytes, year, month, len(aff_sites), aff_matched["_site"].nunique() if not aff_matched.empty else 0)
            st.balloons()
        except Exception as exc:
            st.exception(exc)

    if "result" in st.session_state:
        summary, daily, media, excel_bytes, year, month, plan_sites, matched_sites = st.session_state["result"]
        st.success(f"{year}年{month}月のレポートを作成しました。AFF一致サイト: {matched_sites}/{plan_sites}")
        tab1, tab2, tab3 = st.tabs(["合算", "日別", "媒体別"])
        with tab1:
            display_summary(summary)
        with tab2:
            st.dataframe(daily, use_container_width=True, hide_index=True, height=620)
        with tab3:
            st.dataframe(media, use_container_width=True, hide_index=True, height=620)
        st.download_button(
            "統合レポートをExcelでダウンロード",
            data=excel_bytes,
            file_name=f"統合レポート_{year}{month:02d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )


if __name__ == "__main__":
    main()
